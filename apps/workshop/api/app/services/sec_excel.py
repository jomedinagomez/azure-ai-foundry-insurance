"""Excel export for SEC financial-tables analyzer results.

Ported verbatim from SECExtraction/scripts/export_to_excel.py and made
import-friendly (no CLI side effects on import). Both the service layer
(sec_service.export_to_excel) and notebooks
(demo/sec/notebooks/04_merge_and_export.ipynb) call into this module so
there is exactly one Excel rendering path.

Each financial table becomes its own sheet:
  - Row 1: merged title (tableTitle + companyName + "(in <unit>)"), bold, 14pt.
  - Row 2: spanning group headers (when present), italic, merged consecutively.
  - Row 3: leaf column headers, bold, bottom border.
  - Rows 4+: lineItem with " "*3*level indentation; section headers bold,
            subtotals bold + top border.

Handles both nested (periodHeaders[]/values[]) and portal-flat
(periodHeader1..6/value1..6) CU response shapes, and applies known
CU-ordering fixups for displaced group headers and child rows.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MAX_FLAT_COLS = 6
SHEET_NAME_MAX = 31
FORBIDDEN_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
ILLEGAL_CELL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(s: Any) -> Any:
    if isinstance(s, str):
        return ILLEGAL_CELL_CHARS.sub("", s)
    return s


def _scalar(field: dict[str, Any] | None, default: Any = "") -> Any:
    if not field:
        return default
    for key in ("valueString", "valueInteger", "valueNumber", "valueBoolean"):
        if key in field:
            return _clean(field[key])
    return default


def _array(field: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not field:
        return []
    return field.get("valueArray", []) or []


def _obj(field: dict[str, Any] | None) -> dict[str, Any]:
    if not field:
        return {}
    return field.get("valueObject", {}) or {}


def _conf(field: dict[str, Any] | None) -> float | None:
    """Return the field's `confidence` value if present (0.0–1.0), else None.

    CU emits `.confidence` alongside scalar values for extracted table fields;
    we propagate it so downstream consumers can render review heat-maps.
    """
    if not field:
        return None
    c = field.get("confidence")
    if c is None:
        return None
    try:
        return float(c)
    except (TypeError, ValueError):
        return None


def _normalize_table(raw_tbl: dict[str, Any]) -> dict[str, Any]:
    tbl_obj = _obj(raw_tbl)
    title = str(_scalar(tbl_obj.get("tableTitle"))).strip()
    company = str(_scalar(tbl_obj.get("companyName"))).strip()
    stype = str(_scalar(tbl_obj.get("statementType"))).strip() or "Other"
    unit = str(_scalar(tbl_obj.get("unit"))).strip()

    period_headers: list[str] = []
    period_groups: list[str] = []
    raw_group_count = 0
    if "periodHeader1" in tbl_obj:
        for i in range(1, MAX_FLAT_COLS + 1):
            h = str(_scalar(tbl_obj.get(f"periodHeader{i}"))).strip()
            g = str(_scalar(tbl_obj.get(f"periodGroup{i}"))).strip()
            period_headers.append(h)
            period_groups.append(g)
        while period_headers and not period_headers[-1] and not period_groups[-1]:
            period_headers.pop()
            period_groups.pop()
        raw_group_count = len(period_groups)
    else:
        period_headers = [
            str(_scalar(h)).strip() for h in _array(tbl_obj.get("periodHeaders"))
        ]
        period_groups = [
            str(_scalar(g)).strip()
            for g in _array(tbl_obj.get("periodGroupHeaders"))
        ]
        raw_group_count = len(period_groups)
        if len(period_groups) < len(period_headers):
            period_groups += [""] * (len(period_headers) - len(period_groups))

    rows_norm: list[dict[str, Any]] = []
    for raw_row in _array(tbl_obj.get("rows")):
        row_obj = _obj(raw_row)
        line_item = str(_scalar(row_obj.get("lineItem"))).strip()
        if not line_item:
            continue
        level_val = _scalar(row_obj.get("level"), 0)
        try:
            level = int(level_val) if level_val != "" else 0
        except (TypeError, ValueError):
            level = 0

        if "isSectionHeader" in row_obj:
            _hdr_val = _scalar(row_obj.get("isSectionHeader"), False)
            is_header = (
                str(_hdr_val).lower().strip() == "true"
                if isinstance(_hdr_val, str)
                else bool(_hdr_val)
            )
        else:
            all_empty = all(
                not str(_scalar(v)).strip() or str(_scalar(v)).strip() == "—"
                for v in _array(row_obj.get("values"))
            )
            is_header = level == 0 and (line_item.rstrip().endswith(":") or all_empty)

        if "isSubtotal" in row_obj:
            _sub_val = _scalar(row_obj.get("isSubtotal"), False)
            is_subtotal = (
                str(_sub_val).lower().strip() == "true"
                if isinstance(_sub_val, str)
                else bool(_sub_val)
            )
        else:
            _li_lower = line_item.lower().lstrip()
            is_subtotal = level == 0 and (
                _li_lower.startswith("total ")
                or _li_lower.startswith("net ")
                or _li_lower.startswith("balance,")
                or _li_lower.startswith("balance at")
                or _li_lower.startswith("balance as")
                or _li_lower == "total"
            )

        parent = str(_scalar(row_obj.get("parentLineItem"))).strip()

        if "value1" in row_obj:
            value_fields = [row_obj.get(f"value{i}") for i in range(1, len(period_headers) + 1)]
        else:
            value_fields = list(_array(row_obj.get("values")))
        values = [str(_scalar(f)).strip() for f in value_fields]
        value_confidences = [_conf(f) for f in value_fields]
        # Row-level aggregate = min of available cell confidences (worst-cell wins).
        _present = [c for c in value_confidences if c is not None]
        row_conf = min(_present) if _present else None

        rows_norm.append({
            "lineItem": line_item,
            "level": max(level, 0),
            "parentLineItem": parent,
            "isSectionHeader": is_header,
            "isSubtotal": is_subtotal,
            "values": values,
            "valueConfidences": value_confidences,
            "confidence": row_conf,
        })

    result = {
        "tableTitle": title,
        "companyName": company,
        "statementType": stype,
        "unit": unit,
        "periodGroups": period_groups,
        "periodHeaders": period_headers,
        "rows": rows_norm,
    }
    _fix_header_order(result, raw_group_count)
    _fix_row_order(result)
    return result


def _fix_header_order(table: dict[str, Any], raw_group_count: int) -> None:
    hdrs = table["periodHeaders"]
    n_hdrs = len(hdrs)
    if raw_group_count == 0 or raw_group_count >= n_hdrs:
        return
    grps = table["periodGroups"]
    if not all(g.strip() for g in grps[:raw_group_count]):
        return
    grouped_hdrs = hdrs[-raw_group_count:]
    ungrouped_hdrs = hdrs[:-raw_group_count]
    new_hdrs = grouped_hdrs + ungrouped_hdrs
    new_grps = grps[:raw_group_count] + [""] * len(ungrouped_hdrs)
    table["periodHeaders"] = new_hdrs
    table["periodGroups"] = new_grps


def _fix_row_order(table: dict[str, Any]) -> None:
    rows = table["rows"]
    if len(rows) < 3:
        return
    header_children: dict[str, list[int]] = {}
    header_indices: dict[str, int] = {}
    for i, row in enumerate(rows):
        if row["isSectionHeader"]:
            header_children[row["lineItem"]] = []
            header_indices[row["lineItem"]] = i
        parent = row.get("parentLineItem", "")
        if parent and parent in header_children:
            header_children[parent].append(i)
    needs_fix = False
    for hdr_name, child_idxs in header_children.items():
        if not child_idxs:
            continue
        hdr_idx = header_indices[hdr_name]
        expected_start = hdr_idx + 1
        for offset, ci in enumerate(child_idxs):
            if ci != expected_start + offset:
                needs_fix = True
                break
        if needs_fix:
            break
    if not needs_fix:
        return
    placed: set[int] = set()
    new_rows: list[dict[str, Any]] = []

    def _place_row(idx: int) -> None:
        if idx in placed:
            return
        placed.add(idx)
        row = rows[idx]
        new_rows.append(row)
        if row["isSectionHeader"] and row["lineItem"] in header_children:
            for ci in header_children[row["lineItem"]]:
                _place_row(ci)

    for i in range(len(rows)):
        _place_row(i)
    table["rows"] = new_rows


def load_document(json_path: Path) -> list[dict[str, Any]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return load_from_payload(data)


def load_from_payload(data: dict[str, Any]) -> list[dict[str, Any]]:
    """Same as load_document but accepts an in-memory CU payload."""
    contents = data.get("contents") or []
    if not contents:
        return []
    fields = contents[0].get("fields") or {}
    tables_field = fields.get("financialTables")
    if not tables_field:
        return []
    return [_normalize_table(t) for t in _array(tables_field)]


def _slug(text: str, limit: int) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    return s[:limit] or "table"


def sheet_name_for_table(idx: int, table: dict[str, Any], used: set[str]) -> str:
    stype = table["statementType"] or "Other"
    prefix = f"{idx:02d}_{stype}_"
    slug_budget = SHEET_NAME_MAX - len(prefix)
    base = prefix + _slug(table["tableTitle"] or stype, slug_budget)
    base = FORBIDDEN_SHEET_CHARS.sub("", base)[:SHEET_NAME_MAX]
    name = base
    n = 2
    while name in used:
        suffix = f"_{n}"
        name = base[: SHEET_NAME_MAX - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


_THIN = Side(style="thin", color="000000")
_BORDER_BOTTOM = Border(bottom=_THIN)
_BORDER_TOP = Border(top=_THIN)
_FONT_TITLE = Font(bold=True, size=14)
_FONT_GROUP = Font(italic=True, size=10)
_FONT_HEADER = Font(bold=True, size=11)
_FONT_SECTION = Font(bold=True)
_FONT_SUBTOTAL = Font(bold=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def _merge_consecutive(values: list[str]) -> list[tuple[int, int, str]]:
    spans: list[tuple[int, int, str]] = []
    i = 0
    while i < len(values):
        v = values[i]
        j = i
        while j + 1 < len(values) and values[j + 1] == v:
            j += 1
        spans.append((i, j, v))
        i = j + 1
    return spans


def write_table_sheet(ws: Worksheet, table: dict[str, Any]) -> None:
    headers = table["periodHeaders"]
    groups = table["periodGroups"]
    n_cols = len(headers)
    total_cols = 1 + n_cols

    bits = [b for b in (table["tableTitle"], table["companyName"]) if b]
    title_text = " — ".join(bits) if bits else (table["statementType"] or "Table")
    if table["unit"]:
        title_text += f"  (in {table['unit']})"
    ws.cell(row=1, column=1, value=title_text).font = _FONT_TITLE
    ws.cell(row=1, column=1).alignment = _ALIGN_LEFT
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 22

    data_start_row = 3
    has_groups = any(g for g in groups)
    if has_groups:
        spans = _merge_consecutive(groups)
        for start, end, grp in spans:
            if not grp:
                continue
            c = ws.cell(row=2, column=2 + start, value=grp)
            c.font = _FONT_GROUP
            c.alignment = _ALIGN_CENTER
            if end > start:
                ws.merge_cells(
                    start_row=2, start_column=2 + start,
                    end_row=2, end_column=2 + end,
                )
        data_start_row = 4

    header_row = data_start_row - 1
    c = ws.cell(row=header_row, column=1, value="Line Item")
    c.font = _FONT_HEADER
    c.alignment = _ALIGN_LEFT
    c.border = _BORDER_BOTTOM
    for i, h in enumerate(headers):
        c = ws.cell(row=header_row, column=2 + i, value=h)
        c.font = _FONT_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_BOTTOM

    r = data_start_row
    for row in table["rows"]:
        indent = "   " * row["level"]
        label_cell = ws.cell(row=r, column=1, value=f"{indent}{row['lineItem']}")
        label_cell.alignment = _ALIGN_LEFT
        if row["isSectionHeader"]:
            label_cell.font = _FONT_SECTION
        elif row["isSubtotal"]:
            label_cell.font = _FONT_SUBTOTAL
            label_cell.border = _BORDER_TOP

        vals = row["values"]
        for i in range(n_cols):
            v = vals[i] if i < len(vals) else ""
            cell = ws.cell(row=r, column=2 + i, value=v)
            cell.alignment = _ALIGN_RIGHT
            if row["isSubtotal"]:
                cell.font = _FONT_SUBTOTAL
                cell.border = _BORDER_TOP
            elif row["isSectionHeader"]:
                cell.font = _FONT_SECTION
        r += 1

    ws.freeze_panes = ws.cell(row=data_start_row, column=2)
    ws.column_dimensions["A"].width = 60
    for i in range(n_cols):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18


def export_document(json_path: Path, out_dir: Path) -> Path:
    tables = load_document(json_path)
    out_path = out_dir / f"{json_path.stem}.xlsx"
    out_dir.mkdir(parents=True, exist_ok=True)
    _write_workbook(tables, out_path, fallback_label=json_path.name)
    return out_path


def export_payload(payload: dict[str, Any], out_path: Path) -> Path:
    """Render an in-memory merged CU payload to an .xlsx at `out_path`."""
    tables = load_from_payload(payload)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    _write_workbook(tables, out_path, fallback_label=out_path.name)
    return out_path


def _write_workbook(tables: list[dict[str, Any]], out_path: Path, fallback_label: str) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    if not tables:
        ws = wb.create_sheet(title="empty")
        ws.cell(row=1, column=1, value=f"No financialTables found in {fallback_label}")
    else:
        used: set[str] = set()
        for i, tbl in enumerate(tables, start=1):
            name = sheet_name_for_table(i, tbl, used)
            ws = wb.create_sheet(title=name)
            write_table_sheet(ws, tbl)
    wb.save(out_path)
