"""xlsx page-setup preflight for clean PDF rendering.

The `apply_print_preflight` function mutates an xlsx so that any downstream
PDF engine (LibreOffice, Excel, Aspose) produces a 1-page-wide layout that
preserves anchored images and uses column widths sized to fit every value.

This is the single most important variable in the xlsx→PDF pipeline. Without
it, default Excel print rendering wraps wide schedules across many pages with
mid-row column breaks — and the rendered PDF is unreadable for both humans
and CU's layout analyzer.

See [research_xlsx_to_pdf.ipynb](../../feedback/underwriting/research_xlsx_to_pdf.ipynb)
for the matrix that established these defaults.
"""
from __future__ import annotations

import math
from pathlib import Path

_EMU_PER_PX = 9525           # English Metric Unit constant
_DEFAULT_ROW_PX = 20          # ~15pt
_DEFAULT_COL_PX = 64          # ~8.43 char width @ Calibri 11pt


def _image_extent_cells(image) -> tuple[int | None, int | None]:
    """Approximate the (end_row_1based, end_col_1based) extent of an anchored
    image so we can extend the print area to cover it.

    Without this, `ws.max_row`/`ws.max_column` only count cell-grid content
    and anchored images get cropped off the printout. The Acme SOV has an
    "ADDITIONAL LOCATIONS" image at row 45 that holds 3 extra rows of the
    schedule; this function ensures it survives the conversion.
    """
    anc = image.anchor
    try:
        from_col = anc._from.col + 1
        from_row = anc._from.row + 1
    except Exception:
        return (None, None)
    cx = cy = None
    ext = getattr(anc, "ext", None)
    if ext is not None and getattr(ext, "cx", None) and getattr(ext, "cy", None):
        cx, cy = ext.cx, ext.cy
    to = getattr(anc, "to", None)
    if to is not None:
        try:
            return (to.row + 1, to.col + 1)
        except Exception:
            pass
    if cx and cy:
        cols = max(1, math.ceil(cx / _EMU_PER_PX / _DEFAULT_COL_PX))
        rows = max(1, math.ceil(cy / _EMU_PER_PX / _DEFAULT_ROW_PX))
        return (from_row + rows, from_col + cols)
    return (from_row + 10, from_col + 5)


def _decimal_places(number_format: str, default: int = 0) -> int:
    """Count digits after the decimal point in an Excel number_format token.

    e.g. "$#,##0.00" -> 2, "0.0%" -> 1, "#,##0" -> 0."""
    if not number_format or "." not in number_format:
        return default
    after = number_format.split(".", 1)[1]
    n = 0
    for ch in after:
        if ch in "0#":
            n += 1
        elif ch in "%;,":
            break
    return n if n > 0 else default


def _format_for_display(value, number_format: str | None) -> str:
    """Approximate Excel's displayed text for a cell so we can size columns
    against the *rendered* width, not the raw `str(value)` width.

    The previous implementation measured `len(str(22400000)) == 8`, but Excel
    actually displays `$22,400,000` (11 chars) under a `"$#,##0"` format. If
    the column is sized for 8 chars, the formatted value overflows and Excel
    paints `###` — which then bakes into the rendered TIFF and breaks CU
    extraction of currency / numeric fields. (Observed on cascade SOV: BPP
    and BI columns showed `###` in the TIFF and came back blank.)"""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        nf = (number_format or "").strip()
        if "%" in nf:
            d = _decimal_places(nf, default=0)
            try:
                return f"{value * 100:.{d}f}%"
            except Exception:
                return f"{value}%"
        wants_currency = "$" in nf
        wants_thousands = wants_currency or "#,##" in nf or ("," in nf and "0" in nf)
        d = _decimal_places(nf, default=0 if isinstance(value, int) else 2)
        try:
            if wants_thousands:
                text = f"{value:,.{d}f}"
            elif d > 0:
                text = f"{value:.{d}f}"
            elif isinstance(value, float) and value.is_integer():
                text = f"{int(value)}"
            else:
                text = f"{value:g}"
        except Exception:
            text = str(value)
        if wants_currency:
            text = "$" + text
        return text
    return str(value)


def _measure_cell(cell) -> int:
    v = cell.value
    if v is None or v == "":
        return 0
    s = _format_for_display(v, getattr(cell, "number_format", None))
    return max((len(line) for line in s.splitlines()), default=0)


def _measure_cell_text(value) -> int:
    if value is None:
        return 0
    s = str(value)
    return max((len(line) for line in s.splitlines()), default=0)


def _default_font_size(wb) -> float:
    """Best-effort lookup of the workbook's default body font size (points).

    Falls back to Calibri 11pt — Excel's classic default — when the
    workbook doesn't pin a default. Used to derive a sensible row height
    instead of hard-coding 20pt for every input.
    """
    try:
        normal = wb.named_styles["Normal"]
        sz = getattr(getattr(normal, "font", None), "size", None)
        if sz:
            return float(sz)
    except Exception:
        pass
    return 11.0


def _adaptive_row_height(wb, *, lo: float = 18.0, hi: float = 24.0) -> float:
    """Derive a uniform row height from the workbook's default font.

    Empirically, `font_size * 1.8` matches the 20pt we proved on 11pt-Calibri
    SOVs (20/11 ≈ 1.82). Clamped to keep dense layouts readable without
    pushing CU's column-detection past its drift threshold (observed at 25pt+
    on coastal).
    """
    size = _default_font_size(wb)
    return max(lo, min(hi, size * 1.8))


def _currency_columns(ws, *, last_row: int) -> set[int]:
    """Return the 1-based column indexes whose cells render with a `$` symbol.

    CU's column-detection heuristic appears to use horizontal whitespace gaps
    between values to separate columns. Currency columns are the densest part
    of an SOV (multi-comma numbers, $-prefix) and are exactly where we've
    seen one-column-right drift when rows get tall. Bumping these columns'
    autofit width by a couple extra chars increases the gap to their
    neighbors and stabilizes column assignment.
    """
    cols: set[int] = set()
    for row in ws.iter_rows(min_row=1, max_row=last_row, values_only=False):
        for cell in row:
            nf = getattr(cell, "number_format", None)
            if nf and "$" in nf:
                cols.add(cell.column)
    return cols


def autofit_columns(
    ws,
    *,
    min_width: float = 8.0,
    max_width: float = 60.0,
    padding: float = 1.25,
    extra_chars: float = 2.0,
) -> None:
    """Approximate Excel's 'AutoFit Column Width', honoring merged-cell ranges.

    openpyxl can't measure font metrics, but `_format_for_display` returns
    the rendered text (e.g. `$22,400,000` instead of `22400000`) so currency
    columns get sized correctly. `padding * width + extra_chars` matches
    Excel's tendency to leave a half-character of slack on either side.
    Merged ranges (e.g. a header spanning A:D) distribute their width across
    the merged columns so one long header doesn't force a single column to
    be huge.
    """
    from openpyxl.utils import get_column_letter

    merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_lookup[(r, c)] = (mr.min_col, mr.max_col)

    col_max: dict[int, float] = {}
    last_row = ws.max_row or 1
    for row_cells in ws.iter_rows(min_row=1, max_row=last_row, values_only=False):
        for cell in row_cells:
            text_w = _measure_cell(cell)
            if not text_w:
                continue
            r, c = cell.row, cell.column
            if (r, c) in merged_lookup:
                cmin, cmax = merged_lookup[(r, c)]
                span = cmax - cmin + 1
                share = max(1.0, text_w / span)
                for cc in range(cmin, cmax + 1):
                    col_max[cc] = max(col_max.get(cc, 0), share)
            else:
                col_max[c] = max(col_max.get(c, 0), text_w)

    for col_idx, width in col_max.items():
        new_w = max(min_width, min(max_width, width * padding + extra_chars))
        ws.column_dimensions[get_column_letter(col_idx)].width = new_w


def reset_row_heights(ws) -> None:
    """Drop explicit row heights so the renderer uses defaults."""
    for rd in ws.row_dimensions.values():
        rd.height = None


def set_row_heights(ws, *, last_row: int, height: float) -> None:
    """Force every printed row to a uniform `height` (points).

    Excel's default row height is ~15 pt. Bumping to ~28 pt roughly
    doubles vertical padding between values, which gives CU's layout
    analyzer cleaner row boundaries and reduces off-by-one row
    misalignment in dense tables.
    """
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].height = height


def _last_data_row(ws) -> int:
    """Return the 1-based index of the last row containing visible content.

    Plain `ws.max_row` often reports a row well past the last data row
    because openpyxl preserves styled-but-empty rows. Printing through that
    row produces a TIFF with a tall blank tail — wasted bytes and pixels for
    CU to process. We walk back from the bottom and stop at the first row
    that has at least one cell with a non-blank value or a merged span.
    """
    merged_max_row = 1
    for mr in ws.merged_cells.ranges:
        if mr.max_row > merged_max_row:
            merged_max_row = mr.max_row
    last = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1, values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            if cell.row > last:
                last = cell.row
            break  # any non-empty cell on this row qualifies the row
    return max(last, merged_max_row)


def _locations_start_row(ws, *, last_row: int) -> int:
    """Heuristic: find the first row of the locations table.

    Account-summary sections are typically narrow (label + value, 2 cells per
    row). Location tables are wide — the header row has many filled cells
    (Loc #, Address, City, State, Construction, ...). We scan downward and
    pick the first row with >= 4 non-empty cells whose values look like
    column labels (any string). Returns `last_row + 1` (i.e. "no table") if
    nothing qualifies — in that case the caller should skip centering.
    """
    for r in range(1, last_row + 1):
        filled = 0
        for cell in ws[r]:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            filled += 1
            if filled >= 4:
                return r
    return last_row + 1


def center_align_cells(ws, *, first_row: int, last_row: int, last_col: int) -> None:
    """Apply horizontal+vertical center alignment to a row range.

    Centering is applied only to the locations-table rows (passed in via
    `first_row`) — account-summary rows above it are left untouched. Centering
    the account section caused CU to mis-bind label/value pairs because the
    two-column layout's column boundary moved when each value was re-centered
    inside its cell. Wrap-text/rotation/indent on existing styles are preserved.
    """
    from openpyxl.styles import Alignment
    if first_row > last_row:
        return
    for row in ws.iter_rows(
        min_row=first_row, max_row=last_row, max_col=last_col, values_only=False
    ):
        for cell in row:
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=existing.wrap_text if existing else False,
                text_rotation=existing.text_rotation if existing else 0,
                indent=existing.indent if existing else 0,
                shrink_to_fit=existing.shrink_to_fit if existing else False,
            )


def apply_print_preflight(
    xlsx_path: str | Path,
    out_path: str | Path | None = None,
    *,
    autofit: bool = True,
    print_gridlines: bool = False,
    trim_blank_rows: bool = True,
    center_cells: bool = True,
    row_height: float | None = None,
) -> Path:
    """Mutate the workbook's page-setup so any PDF engine produces a
    1-page-wide layout that includes anchored images and has columns wide
    enough to show every value.

    Returns the path to the new preflighted xlsx (the input is never mutated).

    `print_gridlines=False` (default) reproduces Excel's natural print output:
    only the borders the workbook itself defines are visible. Set to True if
    you want Excel's standard light gridlines printed everywhere.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    xlsx_path = Path(xlsx_path)
    out_path = (
        Path(out_path)
        if out_path is not None
        else xlsx_path.with_name(xlsx_path.stem + ".print-ready.xlsx")
    )

    wb = load_workbook(xlsx_path)
    effective_row_height = row_height if row_height is not None else _adaptive_row_height(wb)
    for ws in wb.worksheets:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

        ws.print_options.gridLines = print_gridlines
        ws.print_options.gridLinesSet = print_gridlines

        if autofit:
            autofit_columns(ws)
            reset_row_heights(ws)

        last_row = _last_data_row(ws) if trim_blank_rows else (ws.max_row or 1)
        last_col = ws.max_column or 1
        for im in list(ws._images):
            er, ec = _image_extent_cells(im)
            if er:
                last_row = max(last_row, er)
            if ec:
                last_col = max(last_col, ec)

        if center_cells:
            loc_start = _locations_start_row(ws, last_row=last_row)
            center_align_cells(
                ws, first_row=loc_start, last_row=last_row, last_col=last_col
            )

        if row_height is not None or effective_row_height:
            set_row_heights(ws, last_row=last_row, height=effective_row_height)

        ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"


        ws.page_margins = PageMargins(
            left=0.3, right=0.3, top=0.3, bottom=0.3, header=0.2, footer=0.2
        )
        ws.print_title_rows = "1:1"

    wb.save(out_path)
    return out_path
