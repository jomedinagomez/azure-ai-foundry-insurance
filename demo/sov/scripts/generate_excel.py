"""Generate Excel SOV variants 1, 2, 3, 6 — each with a distinct broker template style."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont

from seed_data import ACME, CASCADE, MAGNOLIA, COASTAL, ROOT, Account, Location

ATTACH_DIR = ROOT / "attachments"
TMP_DIR = ROOT / "scripts" / ".tmp"

# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hex(h: str) -> str:
    return h.lstrip("#")


def _header_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=_hex(color_hex))


def _set_col_widths(ws, widths: Iterable[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _font(size: int) -> ImageFont.FreeTypeFont:
    for path in (r"C:\Windows\Fonts\segoeui.ttf", r"C:\Windows\Fonts\arial.ttf"):
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


# --------------------------------------------------------------------------- #
# Variant 1: Acme — header block + table + embedded image
# --------------------------------------------------------------------------- #

def _make_extra_locations_image(locs: list[Location]) -> Path:
    """Render a small PNG that shows 3 extra locations not in the main table — simulates a screenshot the broker pasted in."""
    W, H = 880, 180
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)
    title_font = _font(16)
    h_font = _font(13)
    cell_font = _font(12)

    draw.text((10, 8), "Additional Locations (added via M&A — Q1 2026)", fill="black", font=title_font)
    headers = ["Loc #", "Address", "City", "State", "Bldg Value", "BPP", "BI"]
    col_x =  [10,      80,        320,    400,     460,           590,    700]
    col_w =  [70,      240,       80,     60,      130,           110,    150]

    # Header row
    y = 36
    draw.rectangle([(8, y - 2), (W - 8, y + 22)], fill=(230, 230, 240))
    for x, h in zip(col_x, headers):
        draw.text((x + 4, y + 2), h, fill="black", font=h_font)

    # Data rows
    y = 62
    for loc in locs:
        for x, val in zip(col_x, [
            loc.location_number,
            loc.street,
            loc.city,
            loc.state,
            f"${loc.building_value:,.0f}" if loc.building_value else "",
            f"${loc.bpp_value:,.0f}" if loc.bpp_value else "",
            f"${loc.bi_ee_value:,.0f}" if loc.bi_ee_value else "",
        ]):
            draw.text((x + 4, y + 2), str(val), fill="black", font=cell_font)
        # row underline
        draw.line([(8, y + 22), (W - 8, y + 22)], fill=(220, 220, 220))
        y += 28

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    out = TMP_DIR / "acme_extra_locations.png"
    img.save(out, "PNG")
    return out


def build_acme(acc: Account) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "SOV"

    color = acc.broker.color
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")

    # ---- Header block (label/value pairs) ----
    ws.merge_cells("A1:N1")
    ws["A1"] = f"STATEMENT OF VALUES — {acc.insured_name}"
    ws["A1"].fill = _header_fill(color)
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    header_pairs = [
        ("Named Insured",          acc.insured_name),
        ("DBA",                    acc.dba or ""),
        ("Mailing Address",        acc.mailing_address),
        ("Effective Date",         acc.effective_date),
        ("Policy Period",          f"{acc.effective_date} to {acc.expiration_date}"),
        ("Currency",               acc.currency),
        ("Valuation Date",         acc.valuation_date),
        ("Primary Operations",     acc.primary_operations),
        ("NAICS",                  acc.naics or ""),
        ("Total Insured Value",    f"${sum(l.tiv for l in acc.locations):,.0f}"),
        ("Number of Locations",    f"{len(acc.locations)} (+3 see embedded image below)"),
        ("Broker",                 acc.broker.name),
        ("Producer Contact",       f"{acc.broker.contact} | {acc.broker.email} | {acc.broker.phone}"),
        ("Prepared By / Date",     f"{acc.prepared_by} — {acc.prepared_date}"),
    ]
    row = 3
    for label, value in header_pairs:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="EEEEEE")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # ---- Schedule of Locations ----
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row=row, column=1, value="SCHEDULE OF LOCATIONS").font = title_font
    ws.cell(row=row, column=1).fill = _header_fill(color)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    row += 2

    headers = [
        "Loc #", "Building #", "Street Address", "City", "State", "ZIP",
        "Construction", "Occupancy / Operations", "Year Built", "Sq Ft",
        "Bldg Value", "BPP Value", "BI/EE Value", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = bold_white
        c.fill = _header_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    data_start = row + 1

    # Data rows — exclude locations 20-22 (those go into the embedded image)
    visible_locs = [l for l in acc.locations if l.location_number < 20]
    extra_locs = [l for l in acc.locations if l.location_number >= 20]

    for r_offset, loc in enumerate(visible_locs):
        r = data_start + r_offset
        operations = f"{loc.occupancy} — {loc.operations_description}" if loc.operations_description else (loc.occupancy or "")
        values = [
            loc.location_number, loc.building_number or "", loc.street, loc.city, loc.state, loc.zip or "",
            loc.construction_type or "", operations, loc.year_built or "", loc.square_footage or "",
            loc.building_value or "", loc.bpp_value or "", loc.bi_ee_value or "", loc.notes or "",
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=r, column=col, value=v).border = BORDER
        for col in (11, 12, 13):  # currency columns
            ws.cell(row=r, column=col).number_format = '"$"#,##0'

    # Subtotal row by state — simple grand total
    total_row = data_start + len(visible_locs) + 1
    ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = bold
    ws.cell(row=total_row, column=10, value=sum(l.square_footage or 0 for l in visible_locs)).font = bold
    ws.cell(row=total_row, column=11, value=sum(l.building_value or 0 for l in visible_locs)).font = bold
    ws.cell(row=total_row, column=11).number_format = '"$"#,##0'
    ws.cell(row=total_row, column=12, value=sum(l.bpp_value or 0 for l in visible_locs)).font = bold
    ws.cell(row=total_row, column=12).number_format = '"$"#,##0'
    ws.cell(row=total_row, column=13, value=sum(l.bi_ee_value or 0 for l in visible_locs)).font = bold
    ws.cell(row=total_row, column=13).number_format = '"$"#,##0'
    for col in range(1, 15):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="DDDDDD")

    # Hidden column: insert a "Reserved (do not use)" column G later — emulate hidden col by setting hidden=True
    # We'll simply hide column F (ZIP) for one of the inserted "phantom" cols — instead emulate by hiding column N's neighbor.
    # Add a hidden helper column at end (P) with internal notes.
    ws.cell(row=row, column=15, value="Internal").font = Font(italic=True, color="888888")
    for r_offset, loc in enumerate(visible_locs):
        ws.cell(row=data_start + r_offset, column=15, value="lookup_key_" + str(loc.location_number))
    ws.column_dimensions[get_column_letter(15)].hidden = True

    # Column widths
    _set_col_widths(ws, [7, 10, 28, 16, 7, 8, 18, 38, 10, 10, 14, 14, 14, 28, 14])

    # ---- Embedded image of extra locations ----
    img_row = total_row + 3
    ws.merge_cells(start_row=img_row, start_column=1, end_row=img_row, end_column=14)
    ws.cell(row=img_row, column=1, value="ADDITIONAL LOCATIONS (image — added late in cycle):").font = bold

    img_path = _make_extra_locations_image(extra_locs)
    xl_img = XLImage(str(img_path))
    ws.add_image(xl_img, f"A{img_row + 1}")

    # Freeze header
    ws.freeze_panes = "A19"

    out = ATTACH_DIR / "01_acme_SOV.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
# Variant 2: Cascade — clean flat table (baseline)
# --------------------------------------------------------------------------- #

def build_cascade(acc: Account) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "SOV"
    color = acc.broker.color
    bold_white = Font(bold=True, color="FFFFFF")

    headers = [
        "Loc #", "Address", "City", "State", "ZIP", "Construction", "Occupancy",
        "Year Built", "Square Footage", "Building Value", "BPP", "BI",
        "Sprinklered", "Protection Class",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold_white
        c.fill = _header_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    for i, loc in enumerate(acc.locations, start=2):
        values = [
            loc.location_number, loc.street, loc.city, loc.state, loc.zip,
            loc.construction_type, loc.occupancy, loc.year_built, loc.square_footage,
            loc.building_value, loc.bpp_value, loc.bi_ee_value,
            "Yes" if loc.sprinklered else "No", loc.protection_class,
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=v).border = BORDER
        for col in (10, 11, 12):
            ws.cell(row=i, column=col).number_format = '"$"#,##0'

    _set_col_widths(ws, [7, 28, 16, 7, 8, 18, 18, 11, 14, 16, 14, 14, 12, 10])
    ws.freeze_panes = "A2"

    out = ATTACH_DIR / "02_cascade_SOV.xlsx"
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
# Variant 3: Magnolia — multi-sheet (Summary / Locations / CAT Exposure / Notes)
# --------------------------------------------------------------------------- #

def build_magnolia(acc: Account) -> Path:
    wb = Workbook()
    color = acc.broker.color
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")

    # ---- Summary tab ----
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"SOV SUMMARY — {acc.insured_name}"
    ws["A1"].font = title_font
    ws["A1"].fill = _header_fill(color)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    pairs = [
        ("Named Insured",       acc.insured_name),
        ("DBA",                 acc.dba or ""),
        ("Mailing Address",     acc.mailing_address),
        ("Effective Date",      acc.effective_date),
        ("Currency",            acc.currency),
        ("Valuation Date",      acc.valuation_date),
        ("Primary Operations",  acc.primary_operations),
        ("Total Insured Value", sum(l.tiv for l in acc.locations)),
        ("Location Count",      len(acc.locations)),
        ("Broker",              acc.broker.name),
        ("Producer Contact",    f"{acc.broker.contact} | {acc.broker.email} | {acc.broker.phone}"),
        ("Prepared By",         acc.prepared_by),
        ("Prepared Date",       acc.prepared_date),
    ]
    for r, (label, value) in enumerate(pairs, start=3):
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="EEEEEE")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)) and label == "Total Insured Value":
            c.number_format = '"$"#,##0'
    _set_col_widths(ws, [22, 24, 24, 24, 24, 24])

    # ---- Locations tab ----
    # NOTE: header label "Bldg Type" here (Summary uses "Construction Class" intent in pairs above)
    # Demonstrates label-inconsistency anomaly
    ws2 = wb.create_sheet("Locations")
    headers = [
        "Loc #", "Property Name / Street", "City", "State", "ZIP", "County",
        "Bldg Type",   # <-- demonstrates label inconsistency
        "Use", "Year", "# Stories", "# Units", "Sq Ft",
        "Building $", "Contents $", "Business Income $",
        "Sprinkler", "Prot. Class", "Roof Yr", "Flood Zone", "Dist Coast (mi)", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = bold_white
        c.fill = _header_fill(color)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 32

    for i, loc in enumerate(acc.locations, start=2):
        values = [
            loc.location_number, loc.street, loc.city, loc.state, loc.zip, loc.county,
            loc.construction_type, loc.occupancy, loc.year_built, loc.stories, loc.unit_count, loc.square_footage,
            loc.building_value, loc.bpp_value, loc.bi_ee_value,
            "Yes" if loc.sprinklered else "No", loc.protection_class, loc.roof_year, loc.flood_zone, loc.distance_to_coast_mi,
            loc.notes,
        ]
        for col, v in enumerate(values, start=1):
            ws2.cell(row=i, column=col, value=v).border = BORDER
        for col in (13, 14, 15):
            ws2.cell(row=i, column=col).number_format = '"$"#,##0'

    _set_col_widths(ws2, [7, 28, 14, 7, 8, 14, 22, 22, 8, 9, 9, 10, 14, 13, 16, 10, 10, 9, 11, 13, 30])

    # ---- CAT Exposure tab ----
    ws3 = wb.create_sheet("CAT Exposure")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "CAT EXPOSURE SUMMARY"
    ws3["A1"].font = title_font
    ws3["A1"].fill = _header_fill(color)
    ws3["A1"].alignment = Alignment(horizontal="center")

    cat_headers = ["Loc #", "City, State", "Peril Flags", "Flood Zone", "TIV"]
    for col, h in enumerate(cat_headers, start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = bold_white
        c.fill = _header_fill(color)
        c.border = BORDER
    cat_locs = [l for l in acc.locations if l.cat_zone_flags]
    for i, loc in enumerate(cat_locs, start=4):
        ws3.cell(row=i, column=1, value=loc.location_number).border = BORDER
        ws3.cell(row=i, column=2, value=f"{loc.city}, {loc.state}").border = BORDER
        ws3.cell(row=i, column=3, value=", ".join(loc.cat_zone_flags)).border = BORDER
        ws3.cell(row=i, column=4, value=loc.flood_zone or "").border = BORDER
        c = ws3.cell(row=i, column=5, value=loc.tiv)
        c.number_format = '"$"#,##0'
        c.border = BORDER
    _set_col_widths(ws3, [7, 24, 50, 12, 16])

    # ---- Notes tab ----
    ws4 = wb.create_sheet("Notes")
    ws4["A1"] = "Broker Notes & Annotations"
    ws4["A1"].font = title_font
    ws4["A1"].fill = _header_fill(color)
    ws4.merge_cells("A1:D1")
    ws4["A3"] = (
        "1. Significant Gulf Coast hurricane exposure across LA, MS, FL locations. "
        "Recommend named-storm sublimit review at quote.\n"
        "2. Location 3 (555 Magazine St) recently renovated — please verify replacement cost basis.\n"
        "3. All FL coastal locations carry wind tier 1 designation; storm-surge values per 2024 ARA model.\n"
        "4. Historic structures at locations 1, 11, and 15 — limited retrofit options."
    )
    ws4["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[3].height = 120
    ws4.column_dimensions["A"].width = 90

    out = ATTACH_DIR / "03_magnolia_SOV.xlsx"
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
# Variant 6: Coastal — messy broker template
# --------------------------------------------------------------------------- #

def build_coastal(acc: Account) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Submission"
    color = acc.broker.color
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    italic = Font(italic=True, color="666666")

    # ---- 3-row header block ----
    ws.merge_cells("A1:M1")
    ws["A1"] = "ATLANTIC SPECIALTY BROKERS — Property Submission"
    ws["A1"].font = title_font
    ws["A1"].fill = _header_fill(color)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Insured: {acc.insured_name}    |    Effective: {acc.effective_date}    |    Currency: {acc.currency} (mostly)"
    ws["A2"].font = bold

    ws.merge_cells("A3:M3")
    ws["A3"] = f"Producer: {acc.broker.contact} ({acc.broker.email}, {acc.broker.phone})    Prepared: {acc.prepared_date}"

    # Blank row 4
    # ---- Multi-row column headers (rows 5-6) ----
    # Top row: grouped headers
    ws.merge_cells("A5:A6"); ws["A5"] = "Loc"
    ws.merge_cells("B5:E5"); ws["B5"] = "Property Address"
    ws.merge_cells("F5:F6"); ws["F5"] = "Construction"
    ws.merge_cells("G5:G6"); ws["G5"] = "Occupancy / Use"
    ws.merge_cells("H5:I5"); ws["H5"] = "Building"
    ws.merge_cells("J5:K5"); ws["J5"] = "Contents"
    ws.merge_cells("L5:L6"); ws["L5"] = "Bldg Val"   # <-- LABEL DRIFT (anomaly)
    ws.merge_cells("M5:M6"); ws["M5"] = "Notes"
    # Sub-headers (row 6)
    sub = {"B6": "Street", "C6": "City", "D6": "State", "E6": "ZIP",
           "H6": "Year",   "I6": "Sq Ft",
           "J6": "BPP",    "K6": "BI/EE"}
    for cell, label in sub.items():
        ws[cell] = label
    for cell in ["A5", "B5", "F5", "G5", "H5", "J5", "L5", "M5",
                 "B6", "C6", "D6", "E6", "H6", "I6", "J6", "K6"]:
        ws[cell].font = bold_white
        ws[cell].fill = _header_fill(color)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws[cell].border = BORDER
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 22

    # ---- Data rows starting row 7 ----
    # Insert a mid-table subtotal block, blank row, label-drift example
    r = 7
    us_locs = [l for l in acc.locations if l.state != "NS"]
    cad_locs = [l for l in acc.locations if l.state == "NS"]

    # First write US locations
    for loc in us_locs:
        notes_extra = ""
        # Mid-list, change the column L label-meaning by writing currency in notes
        ws.cell(row=r, column=1, value=loc.location_number).border = BORDER
        ws.cell(row=r, column=2, value=loc.street).border = BORDER
        ws.cell(row=r, column=3, value=loc.city).border = BORDER
        ws.cell(row=r, column=4, value=loc.state).border = BORDER
        ws.cell(row=r, column=5, value=loc.zip).border = BORDER
        ws.cell(row=r, column=6, value=loc.construction_type).border = BORDER
        ws.cell(row=r, column=7, value=f"{loc.occupancy} — {loc.operations_description}").border = BORDER
        ws.cell(row=r, column=8, value=loc.year_built).border = BORDER
        ws.cell(row=r, column=9, value=loc.square_footage).border = BORDER
        ws.cell(row=r, column=10, value=loc.bpp_value).border = BORDER
        ws.cell(row=r, column=10).number_format = '"$"#,##0'
        ws.cell(row=r, column=11, value=loc.bi_ee_value).border = BORDER
        ws.cell(row=r, column=11).number_format = '"$"#,##0'
        ws.cell(row=r, column=12, value=loc.building_value).border = BORDER
        ws.cell(row=r, column=12).number_format = '"$"#,##0'
        ws.cell(row=r, column=13, value=loc.notes or notes_extra).border = BORDER
        r += 1

    # Mid-table SUBTOTAL row (US only) — labelled "RC Bldg" (label-drift #2)
    ws.cell(row=r, column=1, value="US SUBTOTAL").font = bold
    ws.cell(row=r, column=10, value=sum(l.bpp_value or 0 for l in us_locs)).font = bold
    ws.cell(row=r, column=10).number_format = '"$"#,##0'
    ws.cell(row=r, column=11, value=sum(l.bi_ee_value or 0 for l in us_locs)).font = bold
    ws.cell(row=r, column=11).number_format = '"$"#,##0'
    ws.cell(row=r, column=12, value=sum(l.building_value or 0 for l in us_locs)).font = bold
    ws.cell(row=r, column=12).number_format = '"$"#,##0'
    ws.cell(row=r, column=13, value="(values labelled 'RC Bldg' on summary)").font = italic
    for col in range(1, 14):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="EEEEEE")
    r += 1

    # Blank row
    r += 1

    # Section header for Canadian operations
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
    ws.cell(row=r, column=1, value="CANADIAN OPERATIONS — values reported in CAD").font = bold
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FFE9C7")
    r += 1

    for loc in cad_locs:
        ws.cell(row=r, column=1, value=loc.location_number).border = BORDER
        ws.cell(row=r, column=2, value=loc.street).border = BORDER
        ws.cell(row=r, column=3, value=loc.city).border = BORDER
        ws.cell(row=r, column=4, value=loc.state).border = BORDER
        ws.cell(row=r, column=5, value=loc.zip).border = BORDER
        ws.cell(row=r, column=6, value=loc.construction_type).border = BORDER
        ws.cell(row=r, column=7, value=f"{loc.occupancy} — {loc.operations_description}").border = BORDER
        ws.cell(row=r, column=8, value=loc.year_built).border = BORDER
        ws.cell(row=r, column=9, value=loc.square_footage).border = BORDER
        ws.cell(row=r, column=10, value=loc.bpp_value).border = BORDER
        ws.cell(row=r, column=10).number_format = '"C$"#,##0'
        ws.cell(row=r, column=11, value=loc.bi_ee_value).border = BORDER
        ws.cell(row=r, column=11).number_format = '"C$"#,##0'
        ws.cell(row=r, column=12, value=loc.building_value).border = BORDER
        ws.cell(row=r, column=12).number_format = '"C$"#,##0'
        ws.cell(row=r, column=13, value=loc.notes or "").border = BORDER
        r += 1

    # Grand total — uses third label "Building Replacement Cost"
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.cell(row=r, column=1, value="GRAND TOTAL — Building Replacement Cost (USD + CAD as reported, NOT FX-normalized):").font = bold
    ws.cell(row=r, column=12, value=sum(l.building_value or 0 for l in acc.locations)).font = bold
    ws.cell(row=r, column=12).number_format = '"$"#,##0'
    for col in range(1, 14):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="DDDDDD")

    _set_col_widths(ws, [6, 26, 14, 7, 10, 18, 36, 7, 10, 14, 14, 14, 30])

    out = ATTACH_DIR / "06_coastal_SOV.xlsx"
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> None:
    print("Generating Excel SOV variants...")
    for builder, acc in [(build_acme, ACME), (build_cascade, CASCADE),
                         (build_magnolia, MAGNOLIA), (build_coastal, COASTAL)]:
        out = builder(acc)
        print(f"  wrote {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
